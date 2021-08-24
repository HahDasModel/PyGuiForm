import pyautogui as pag
from openpyxl import load_workbook
import time
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import uuid
import win32api



wb = Workbook()                                                                                        	# Стандартная инициализация книги Excel из openpyxl    
    
dest_filename = 'CoTpyDHuku.xlsx'                                                                       # Присваивание переменной, Инициализация новой книги для сохранения (Сохраняется в директории с файлом .py)    
    
wb = load_workbook(filename='CoTpyDHuku.xlsx', read_only=False)										    # Загрузка указанного в filename файла Excel    
    
ws = wb['range names']																					# Присваивание листу наименования																    

ws = wb.active                                                                                          # Присваиваем переменную первому листу Excel    
    
ws.title = "range names"

ws['A1'] = 'Порядковый номер сотрудника'                                                                # Даём колонкам в листе Excel имена
ws['B1'] = 'Имя'
ws['C1'] = 'Фамилия'
ws['D1'] = 'Дата рождения'
ws['E1'] = 'Паспортные данные'
ws['F1'] = 'Должность'
ws['G1'] = 'Отдел'
ws['H1'] = 'id'
ws['I1'] = 'password'



newnumber = 'порядковый номер'							    												# Начало блока функции для ввода нового порядкового номера в блоке программы для внесения в excel
def pag_prompt_new_number(newnumber):
    new_number=pag.prompt('Порядковый номер', 'Порядковый номер')
    newnumber=new_number
    return newnumber

newname = 'пример' 																							# Начало блока функции для ввода имени нового сотрудника в блоке программы для внесения в excel
def pag_prompt_new_name(newname):
    new_name=pag.prompt('Введите имя нового сотрудника', 'Имя')
    newname=new_name
    return newname

newsurname = 'пример' 																						# Начало блока функции для ввода фамилии нового сотрудника в блоке программы для внесения в excel
def pag_prompt_new_surname(newsurname):
    new_surname=pag.prompt('Введите фамилию нового сотрудника', 'Фамилия')
    newsurname=new_surname
    return newsurname

newbirthday = 'пример' 																						# Начало блока функции для ввода даты рождения нового сотрудника в блоке программы для внесения в excel
def pag_prompt_new_birthday(newbirthday):
    new_birthday=pag.prompt('Введите дату рождения нового сотрудника', 'Дата рождения')
    newbirthday=new_birthday
    return newbirthday

newpassportdata = 'пример' 																					# Начало блока функции для ввода паспортных данных нового сотрудника в блоке программы для внесения в excel
def pag_prompt_new_passport_data(newpassportdata):
    new_passport_data=pag.prompt('Введите паспортные данные нового сотрудника', 'Паспортные данные')
    newpassportdata=new_passport_data
    return newpassportdata

newposition = 'пример' 																						# Начало блока функции для ввода должности нового сотрудника в блоке программы для внесения в excel
def pag_prompt_new_position(newposition):
    new_position=pag.prompt('Введите должность нового сотрудника', 'Должность')
    newposition=new_position
    return newposition

newdepartment = 'пример' 																					# Начало блока функции для ввода отдела нового сотрудника в блоке программы для внесения в excel
def pag_prompt_new_department(newdepartment):
    new_department=pag.prompt('Введите отдел нового сотрудника', 'Отдел')
    newdepartment=new_department
    return newdepartment

newpassword = 'пример' 																						# Начало блока функции для ввода пароля нового сотрудника в блоке программы для внесения в excel
def pag_prompt_new_password(newpassword):
    new_password=pag.password('На экране пароль нового сотрудника, можете записать', 'password')
    newpassword = new_password
    return newpassword

a=pag.confirm('Выберите что вы хотите сделать',
        'Управление данными персонала',
        buttons=['Новый сотрудник', 'Удаление данных сотрудника'])

if a =='Новый сотрудник':       # Графа функций, описывающая поле ввода новый сотрудник как последовательность функций с подфункциями, где реализован объектно-ориентированный подход, это называется объектно-ориентированное программирование(ООП), так же это можно назвать древовидной структурой, где каждая тонкая ветвь идет из толстой или интерпретировать аналогично
    
    win32api.MessageBox(0, '  ВНИМАНИЕ!!!   ПЕРЕД РАБОТОЙ С ПРОГРАММОЙ ДОЛЖНЫ БЫТЬ ЗАКРЫТЫ ФАЙЛЫ EXCEL , С КОТОРЫМИ БУДЕТ РАБОТАТЬ ПРОГРАММА', '                    ВНИМАНИЕ!!!')    
    newnumber=pag_prompt_new_number(newnumber)
    newname=pag_prompt_new_name(newname)
    newsurname=pag_prompt_new_surname(newsurname)
    newbirthday=pag_prompt_new_birthday(newbirthday)
    newpassportdata=pag_prompt_new_passport_data(newpassportdata)
    newposition=pag_prompt_new_position(newposition)
    newdepartment=pag_prompt_new_department(newdepartment)
    newid = uuid.uuid1() 
    win32api.MessageBox(0, 'На экране id этого сотрудника, можете записать', str(newid))
    newpassword=pag_prompt_new_password(newpassword)        
        
    An = newnumber
    Bn = newname
    Cn = newsurname
    Dn = newbirthday
    En = newpassportdata
    Fn = newposition
    Gn = newdepartment
    Hn = newid
    In = newpassword    
        
    Spisok = tuple
    Spisok = str(An), str(Bn), str(Cn), str(Dn), str(En), str(Fn), str(Gn), str(Hn), str(In)    
        
    print(Spisok)    
        
    ws.append(Spisok)    
    


elif a=='Удаление данных сотрудника':

    Y=pag.prompt('Введите строку с данными сотрудника')
    
    K="A"
    X=K+Y
    del ws[X]

    K="B"
    X=K+Y
    del ws[X]

    K="C"
    X=K+Y
    del ws[X]

    K="D"
    X=K+Y
    del ws[X]

    K="E"
    X=K+Y
    del ws[X]

    K="F"
    X=K+Y
    del ws[X]

    K="G"
    X=K+Y
    del ws[X]

    K="H"
    X=K+Y
    del ws[X]

    K="I"
    X=K+Y
    del ws[X]



else: 
    pag.alert(text='Программа завершилась с неизвестной ошибкой, при возникновении вопросов обратитесь к создателю', title='Ошибка', button='OK')



example=None
def delete_unusable(example):
    index_row = []

    # loop each row in column A
    for i in range(1, ws.max_row):
        # define emptiness of cell
        if ws.cell(i, 1).value is None:
            # collect indexes of rows
            index_row.append(i)

    # loop each index value
    for row_del in range(len(index_row)):
        ws.delete_rows(idx=index_row[row_del], amount=1)
        # exclude offset of rows through each iteration
        index_row = list(map(lambda k: k - 1, index_row))               # Конец блока

delete_unusable(example)




wb.save(filename = dest_filename) #Сохраняем данные в нашу книгу Excel
wb.close()

exit()
